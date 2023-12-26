import pandas as pd
import sqlite3
from openpyxl import Workbook, load_workbook

def main():
    # SQLite3 database file path
    db_file = 'database\REFUSS_V8.db'

    # SQL query to retrieve data
    queryA = "SELECT CriteriaID, MetricSubID, MetricID, Source, MetricName, MetricQuestion, Answer_Type FROM Metrics;"
    queryB = "SELECT MetricID, Opt1, Opt2, Opt3, Opt4, Opt5, Opt6, Opt7, Opt8, Opt9, Opt10 FROM MetricOptions;"

    # Number of times to reproduce the layout
    n = 0

    try:
        # Establish a connection to the SQLite3 database
        connection = sqlite3.connect(db_file)

        # Get a list of all table names in the database
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        table_names = cursor.fetchall()

        # Create a dictionary to store DataFrames for each table
        dataframes = {}

        # Fetch data from each table and store it in the dictionary
        for table_name in table_names:
            query = f"SELECT * FROM {table_name[0]};"
            dataframes[table_name[0]] = pd.read_sql(query, connection)

        # Close the database connection
        connection.close()

    except Exception as e:
        print("Error connecting to the database:", e)
        exit(1)

    # Create a new Excel workbook
    workbook = Workbook()

    # Get the tables as dataframes from the dictionary
    Dimensions = dataframes["Dimensions"]
    Objectives = dataframes["Objectives"]
    Criteria = dataframes["Criteria"].sort_values(by="CriteriaID", ascending=True)
    Metrics = dataframes["Metrics"].sort_values(by="MetricID", ascending=True)
    MetricsOptions = dataframes["MetricsOptions"].sort_values(by="MetricID", ascending=True)

    # Do some labeling on the dataframes
    Dimensions['ShortLabel'] = Dimensions['DimensionID'].astype(int)
    Dimensions['FullLabel'] = Dimensions['DimensionName']


    for index, row in Objectives.iterrows():
        if row['DimensionID'] == 1:
            DimLetter = 'F'
        else:
            DimLetter = 'P'
        ShortLabel = f"{DimLetter}{row['ObjectiveSubID']}"
        FullLabel = f"{ShortLabel} - {row['ObjectiveName']}"
        Objectives.at[index, 'ShortLabel'] = ShortLabel
        Objectives.at[index, 'FullLabel'] = FullLabel

    for index, row in Criteria.iterrows():
        if row['CriteriaID'].split('.')[0] == '1':
            DimLetter = 'F'
        else:
            DimLetter = 'P'
        ShortLabel = f"{DimLetter}{row['CriteriaID'][2:]}"
        FullLabel = f"{ShortLabel} - {row['CriteriaName']}"
        Criteria.at[index, 'ShortLabel'] = ShortLabel
        Criteria.at[index, 'FullLabel'] = FullLabel
        
    for index, row in Metrics.iterrows():
        if row['MetricID'].split('.')[0] == '1':
            DimLetter = 'F'
        else:
            DimLetter = 'P'
        ShortLabel = f"{DimLetter}{row['MetricID'][2:]}"
        FullLabel = f"{ShortLabel} - {row['MetricName']}"
        Metrics.at[index, 'ShortLabel'] = ShortLabel
        Metrics.at[index, 'FullLabel'] = FullLabel        

    #Initialize empty list to control the need to create new Excel Sheet
    ObejctivesID = []
    CriteriaID = []

    sheet = None
    row_number = 1

    # Iterate through each unique "MetricID" value
    for index, row in Metrics.iterrows():
        
        #Get the values for the metric
        metric_id = row["MetricID"]
        CriterionID = row['CriteriaID']
        
        ObjectiveID = Criteria.loc[Criteria['CriteriaID'] == CriterionID, 'ObjectiveID'].astype(str).item()
        
        DimensionID = Objectives.loc[Objectives['ObjectiveID'] == ObjectiveID, 'DimensionID'].item()
        
        CriteriaShortLabel = Criteria.loc[Criteria['CriteriaID'] == CriterionID, 'ShortLabel'].item()
        CriteriaFullLabel = Criteria.loc[Criteria['CriteriaID'] == CriterionID, 'FullLabel'].item()
        
        ObjectiveFullLabel = Objectives.loc[Objectives['ObjectiveID'] == ObjectiveID, 'FullLabel'].item()
        
        DimensionFullLabel = Dimensions.loc[Dimensions['DimensionID'] == DimensionID, 'FullLabel'].item()
        
        AnswersOptions = MetricsOptions.loc[MetricsOptions['MetricID'] == metric_id].values.tolist()
        
        if AnswersOptions:
            AnswersOptions = AnswersOptions[0][1:]
        else:
            AnswersOptions = []
            
        if CriterionID not in CriteriaID:
            sheet = workbook.create_sheet(title=f'{CriteriaShortLabel}')
            CriteriaID.append(CriterionID)
            row_number = 1 # Reset row number when creating a new sheet
            row_number = writeline(sheet, row_number, '', DimensionFullLabel)
            row_number = writeline(sheet, row_number, ObjectiveFullLabel, 'Objective')
            row_number = writeline(sheet, row_number, CriteriaFullLabel, 'Criteria')
            #row_number = writeline(sheet, row_number, '', '')
            
        row_number = writeline(sheet, row_number, row['FullLabel'], 'Metric')
        row_number = writeline(sheet, row_number, row['Source'], 'Source')
        row_number = writeline(sheet, row_number, row['AnswerType'], 'Answer type')
        row_number = writeline(sheet, row_number, row['MetricQuestion'], 'Question')
        
        for i, option in enumerate(AnswersOptions):
            if option:
                row_number = writeline(sheet, row_number, option, f"{i+1})")
        
        row_number = writeline(sheet, row_number, '', '')

    # Remove the default sheet named "Sheet"
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    # Save the final workbook
    output_file = 'database\output_file2.xlsx'
    workbook.save(output_file)
    
def writeline(sheet, row_number, value, label, num_lines=1):
    """
    Writes multiple lines to the specified sheet starting from the given row number with the provided value and label.
    Returns the updated row number after incrementing by the number of lines written.
    """
    for i in range(num_lines):
        sheet[f"B{row_number}"].value = label
        sheet[f"C{row_number}"].value = value
        row_number += 1

    return row_number

if __name__ == '__main__':
    main()
    print('Finished')